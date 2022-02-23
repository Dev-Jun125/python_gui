from tkinter import *
import random
import openpyxl as xl

def random1():
    result = []
    enter_min_result = int((enter_min.get()))
    enter_max_result = int((enter_max.get()))
    enter_amount_result = int((enter_amount.get()))
    for i in range(0,enter_amount_result):
        var = random.randrange(enter_min_result, enter_max_result)
        result.append(var)
    print(result)

    result_lab = Label(window, text = result)
    result_lab.pack()
    
    
    wb = xl.Workbook()
    sheet = wb.active
    sheet.title = '테스트'  
    # 컬럼명 지정(헤더) 
    col_names = ['test'] 
    for seq, name in enumerate(col_names): 
        sheet.cell(row=1, column=seq+1, value=name) 
    # 시트 저장 
    
    row_no = 1 
    for i in range(len(result)):
        sheet.cell(row=row_no+i+1, column=seq+1).value=result[i]
            
        
    wb.save("test.xlsx") 
    wb.close()


window=Tk()

Label(window, text = '최소값').pack(pady=5)
enter_min = Entry(window)
enter_min.pack(pady=10)
enter_min.focus()

Label(window, text = '최대값').pack(pady=5)
enter_max = Entry(window)
enter_max.pack(pady=10)

Label(window, text = '원하는 갯수').pack(pady=5)
enter_amount = Entry(window)
enter_amount.pack(pady=10)


B = Button(text="실행", command=random1)
B.pack(side=RIGHT)

mainloop()

